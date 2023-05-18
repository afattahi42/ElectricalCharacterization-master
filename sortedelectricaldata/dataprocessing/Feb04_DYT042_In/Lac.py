import time
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import matplotlib.pyplot as plt
from random import randint
import pandas as pd
import pyvisa as visa

def main(timeRunning = 5, timeBetweenMeasurement = 1): 
    wb, title = makeSheet() # create sheet
    row = 1
    numMeasurements = timeRunning/timeBetweenMeasurement
    while row < numMeasurements: # Process is running
        oldTime = time.time()
        row+=1
        tcoTemp = outputCall()
        resistance = 0
        resistanceUncertainty = 0
        dataToExcel(title, resistance, resistanceUncertainty, tcoTemp, row)
        newTime = time.time()
        timeDiff = newTime-oldTime
        if timeDiff > timeBetweenMeasurement:
            pass
        else:
            time.sleep(timeBetweenMeasurement - timeDiff)

        
    graphResults(title, row) # graph result




def setParameters():
    rm = visa.ResourceManager()
    instrument = rm.list_resources()[0]
    inst = rm.open_resource(instrument)

def makeSheet():
    wb = Workbook()
    title = str(datetime.fromtimestamp(time.time()))
    newTitle = ''
    for i in title:
        if i == ":":
            newTitle += " "
        else:
            newTitle += i

    title = newTitle+".xlsx"

    ws = wb.create_sheet("Data")
    row = 1

    ws.cell(row=row, column = 1).value = "Time"
    ws.cell(row=row, column = 2).value = "Unix Timestamp"
    ws.cell(row=row, column = 3).value = "Resistance"
    ws.cell(row=row, column =4).value = "Resistance Uncertainty"
    ws.cell(row=row, column =5).value = "TCO Temp"
    wb.save(title)

    return wb,title

def outputCall():
    tcoTemp = thermalCoupleTemp()
    return tcoTemp

def thermalCoupleTemp():
    rm = visa.ResourceManager()
    instrument = rm.list_resources()[0]
    inst = rm.open_resource(instrument)
    ## query for data
    temp = inst.query('meas:tco?')


    return temp
        
def dataToExcel(title, resistance, resistanceUncertainty, tcoTemp, row):
    wb = load_workbook(title)
    ws = wb["Data"]
  
    ws.cell(row = row, column = 1).value = str(datetime.fromtimestamp(time.time()))
    ws.cell(row = row, column = 2).value = (time.time())
    ws.cell(row = row, column = 3).value = resistance
    ws.cell(row = row, column = 4).value = resistanceUncertainty
    ws.cell(row = row, column = 5).value = tcoTemp
    wb.save(title)

def graphResults(title, row):
    wb = load_workbook(title)
    ws = wb["Data"]
    time = []
    resistance = []
    resistanceUncertainty = []

   
    for i in range(row-1):
        i += 2
        time += [ws.cell(row = i, column = 2).value]
        resistance += [ws.cell(row = i, column = 3).value]
        resistanceUncertainty += [ws.cell(row = i, column = 4).value]
    
    xlim = [time[0], time[-1]]
    plt.errorbar(time, resistance, yerr= resistanceUncertainty)
    plt.xlim(xlim)
    plt.savefig(title + ".png", dpi = 150)